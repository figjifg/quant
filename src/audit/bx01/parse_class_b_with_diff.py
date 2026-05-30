#!/usr/bin/env python3
"""BX01-attachment-parse: Class B parse — snapshot table + consecutive-snapshot diff.

User-confirmed cycle mapping via filename rename (e.g., `23.6.xlsx` = 2023-06 review).
Per Referee 2026-05-30 clarification:
- High-confidence mapping (filename direct cycle match) → snapshot table allowed.
- Consecutive-snapshot diff allowed ONLY if BOTH snapshots are high-confidence + KRX
  official + KOSPI200 list clear. All derived rows tagged
  `derived_via_consecutive_official_snapshot_diff`.
- effective_dt remains direct-from-file-only (NOT filled from snapshot or rebalance).
- Bridge-only snapshots (2022-12, NOT in Tier 1) used only to disambiguate consecutive
  diffs; their derived events are flagged `bridge_only_not_in_tier1`.
- Preserve-all + caveats + cross-check note for downstream.

CRITICAL CAVEAT (added explicitly to every derived row):
Snapshot diff between cycle N-1 and cycle N captures ALL constituent changes between
those two snapshots — REGULAR review changes AT cycle N PLUS any SPECIAL/SUPPLEMENTAL
events between. Tier 2/3 specials were DEFERRED in this phase; therefore diff-derived
rows may conflate regular and intermediate-special events. The downstream user
decision must separate them.
"""
from __future__ import annotations

import csv
import hashlib
import re
from pathlib import Path

import openpyxl

SRC = Path("/home/jin/code/quant/research_input_data/정기변경")
OUT_DIR = Path("/home/jin/code/quant/data/acquired/bx01_kospi200_index_event_source_a0")
OUT_SNAPSHOTS = OUT_DIR / "snapshots.csv"
OUT_DIFF_EVENTS = OUT_DIR / "events_class_b_derived.csv"

# User-confirmed mapping: filename → (bbsSeq, reg_dt, cycle, is_tier1)
# Tier 1 bbsSeq from tier1_download_list.csv. 22-12 = bridge (not in Tier 1).
CYCLE_FILES = [
    # (filename, cycle_label, bbsSeq, reg_dt, is_tier1, source_class_marker)
    ("21.6.xlsx",  "2021-06", "734",  "2021-05-25", True,  "B_snapshot_with_change_detail"),  # also Class A; we use snapshot here for diff base
    ("21.12.xlsx", "2021-12", "774",  "2021-11-24", True,  "B_snapshot_only"),
    ("22.6.xlsx",  "2022-06", "799",  "2022-05-24", True,  "B_snapshot_only"),
    ("22.12.xlsx", "2022-12", "",      "",           False, "B_snapshot_only_bridge_not_in_tier1"),
    ("23.6.xlsx",  "2023-06", "861",  "2023-05-18", True,  "B_snapshot_only"),
    ("23.12.xlsx", "2023-12", "899",  "2023-11-23", True,  "B_snapshot_only"),
    ("24.6.xlsx",  "2024-06", "939",  "2024-05-24", True,  "B_snapshot_only"),
    ("24.12.xlsx", "2024-12", "979",  "2024-11-21", True,  "B_snapshot_only"),
    ("25.6.xlsx",  "2025-06", "1015", "2025-05-27", True,  "B_snapshot_only"),
    ("25.12.xlsx", "2025-12", "1050", "2025-11-18", True,  "B_snapshot_only"),
    ("26.6.xlsx",  "2026-06", "1090", "2026-05-22", True,  "B_snapshot_only"),
]


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def normalize_ticker(s) -> str | None:
    if s is None:
        return None
    s = str(s).strip()
    m = re.match(r"^KR7(\d{6})\d{3}$", s)
    if m:
        return m.group(1)
    m = re.match(r"^(\d{6})$", s)
    if m:
        return m.group(1)
    m = re.match(r"^(\d{1,6})$", s)
    if m:
        return m.group(1).zfill(6)
    return None


def extract_kospi200_snapshot(p: Path) -> dict[str, str]:
    """Return {ticker -> company_name} for KOSPI 200 구성종목 sheet."""
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    out: dict[str, str] = {}
    for sn in wb.sheetnames:
        if "KOSPI 200" in sn and "구성종목" in sn:
            ws = wb[sn]
            for r in ws.iter_rows(values_only=True):
                if not r or not r[0]:
                    continue
                t = normalize_ticker(r[0])
                if not t or not t.isdigit():
                    continue
                name = str(r[1]).strip() if len(r) > 1 and r[1] else ""
                out[t] = name
            break
    wb.close()
    return out


def main() -> int:
    # 1) Read all snapshots
    snapshots: dict[str, tuple[dict[str, str], dict]] = {}
    rows_for_snapshot_table = []
    for fn, cycle, bbsseq, regdt, is_tier1, marker in CYCLE_FILES:
        p = SRC / fn
        if not p.exists():
            print(f"MISSING: {p}")
            continue
        ks = extract_kospi200_snapshot(p)
        sha = sha256_file(p)
        if len(ks) != 200:
            print(f"WARN {fn}: KOSPI200 count = {len(ks)} (expected 200)")
        meta = {
            "cycle": cycle, "bbs_seq": bbsseq, "reg_dt": regdt,
            "is_tier1": is_tier1, "source_class_marker": marker,
            "file": fn, "sha256": sha, "bytes": p.stat().st_size,
        }
        snapshots[cycle] = (ks, meta)
        for t, name in ks.items():
            rows_for_snapshot_table.append({
                "cycle": cycle, "bbs_seq": bbsseq, "reg_dt": regdt, "is_tier1": "yes" if is_tier1 else "no",
                "index_name": "KOSPI 200", "ticker": t, "company_name": name,
                "source_file": fn, "source_file_sha256": sha,
                "source_class": marker,
                "source_record_type": "snapshot_only" if "snapshot_only" in marker else "snapshot_with_change_detail",
                "parse_status": "parsed_snapshot_from_kospi200_구성종목_sheet",
                "caveat": ("user_renamed_filename_authoritative_cycle_per_Referee_high_confidence_mapping; "
                            "snapshot represents post-effective-date constituents for this cycle"),
            })
        print(f"loaded {cycle:9s} (bbsSeq={bbsseq:>4s}, tier1={is_tier1}): {len(ks)} tickers from {fn}")

    # Write snapshots.csv
    cols_snap = list(rows_for_snapshot_table[0].keys())
    with OUT_SNAPSHOTS.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols_snap, lineterminator="\n")
        w.writeheader()
        w.writerows(rows_for_snapshot_table)
    print(f"\nwrote {len(rows_for_snapshot_table)} snapshot rows to {OUT_SNAPSHOTS}")

    # 2) Compute consecutive snapshot diffs
    diff_rows = []
    cycles_in_order = [c for (_,c,*_) in CYCLE_FILES]
    eid_counter = 1
    for i in range(1, len(cycles_in_order)):
        prev_cycle = cycles_in_order[i-1]
        curr_cycle = cycles_in_order[i]
        prev_set, prev_meta = snapshots[prev_cycle]
        curr_set, curr_meta = snapshots[curr_cycle]

        added = set(curr_set) - set(prev_set)
        deleted = set(prev_set) - set(curr_set)

        is_target_tier1 = curr_meta["is_tier1"]
        is_bridge_only = not is_target_tier1

        for t in sorted(added):
            name = curr_set.get(t, "")
            diff_rows.append({
                "event_id": f"BX01-DIFF-{curr_meta['bbs_seq'] or 'BRIDGE'}-{eid_counter:04d}",
                "review_cycle": curr_cycle,
                "announcement_dt": curr_meta["reg_dt"],
                "announcement_time": "",
                "effective_dt": "",  # NOT in snapshot — direct-only rule
                "index_name": "KOSPI 200",
                "side": "addition",
                "ticker": t,
                "company_name": name,
                "source_class": "B_derived_from_consecutive_snapshot_diff",
                "source_id_or_url": (
                    f"KRX MDCINFO005 bbsSeq={curr_meta['bbs_seq']}"
                    if is_target_tier1
                    else f"bridge_snapshot_for_cycle={curr_cycle}_not_in_tier1"
                ),
                "source_file": curr_meta["file"],
                "source_file_sha256": curr_meta["sha256"],
                "prev_snapshot_file": prev_meta["file"],
                "prev_snapshot_sha256": prev_meta["sha256"],
                "prev_snapshot_cycle": prev_cycle,
                "parse_status": "derived_via_consecutive_official_snapshot_diff",
                "tier1_status": "tier1_target" if is_target_tier1 else "bridge_only_not_in_tier1",
                "caveat": (
                    "DERIVED — captures ALL constituent changes between consecutive KRX-official "
                    "post-effective-date snapshots; this includes the regular review at the curr cycle "
                    "PLUS any SPECIAL/SUPPLEMENTAL events between cycles (Tier 2/3, deferred). "
                    "Downstream user decision required to separate regular-review vs intermediate-special. "
                    "effective_dt NOT filled (Referee rule: direct-from-file only). "
                    "company_name from KRX snapshot row; ticker is 6-digit normalized from KR7…002 ISIN."
                ),
            })
            eid_counter += 1

        for t in sorted(deleted):
            name = prev_set.get(t, "")
            diff_rows.append({
                "event_id": f"BX01-DIFF-{curr_meta['bbs_seq'] or 'BRIDGE'}-{eid_counter:04d}",
                "review_cycle": curr_cycle,
                "announcement_dt": curr_meta["reg_dt"],
                "announcement_time": "",
                "effective_dt": "",
                "index_name": "KOSPI 200",
                "side": "deletion",
                "ticker": t,
                "company_name": name,
                "source_class": "B_derived_from_consecutive_snapshot_diff",
                "source_id_or_url": (
                    f"KRX MDCINFO005 bbsSeq={curr_meta['bbs_seq']}"
                    if is_target_tier1
                    else f"bridge_snapshot_for_cycle={curr_cycle}_not_in_tier1"
                ),
                "source_file": curr_meta["file"],
                "source_file_sha256": curr_meta["sha256"],
                "prev_snapshot_file": prev_meta["file"],
                "prev_snapshot_sha256": prev_meta["sha256"],
                "prev_snapshot_cycle": prev_cycle,
                "parse_status": "derived_via_consecutive_official_snapshot_diff",
                "tier1_status": "tier1_target" if is_target_tier1 else "bridge_only_not_in_tier1",
                "caveat": (
                    "DERIVED — captures ALL constituent changes between consecutive KRX-official "
                    "post-effective-date snapshots; this includes the regular review at the curr cycle "
                    "PLUS any SPECIAL/SUPPLEMENTAL events between cycles (Tier 2/3, deferred). "
                    "Downstream user decision required to separate regular-review vs intermediate-special. "
                    "effective_dt NOT filled (Referee rule: direct-from-file only). "
                    "company_name from KRX snapshot row; ticker is 6-digit normalized from KR7…002 ISIN."
                ),
            })
            eid_counter += 1

        print(f"\n  diff {prev_cycle} -> {curr_cycle} (bbsSeq={curr_meta['bbs_seq']:>4s}, tier1={is_target_tier1}): "
              f"+{len(added)} additions, -{len(deleted)} deletions")
        for t in sorted(added):
            print(f"    +  {t} {curr_set[t]}")
        for t in sorted(deleted):
            print(f"    -  {t} {prev_set[t]}")

    cols_diff = list(diff_rows[0].keys())
    with OUT_DIFF_EVENTS.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols_diff, lineterminator="\n")
        w.writeheader()
        w.writerows(diff_rows)
    print(f"\nwrote {len(diff_rows)} derived diff rows to {OUT_DIFF_EVENTS}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
