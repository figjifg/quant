#!/usr/bin/env python3
"""BX01-attachment-parse: build best-guess mapping for Class B snapshot-only files.

For each Class B (snapshot-only) xlsx, fingerprint the KOSPI 200 stock set against
known marker stocks (KRX-public listing / inclusion events) and bracket the
possible review cycle. Present a per-file proposed mapping for USER CONFIRMATION
per Referee clarification.

OUTPUT: class_b_mapping_proposal.csv with one row per Class B file:
  - filename, sha256, size
  - candidate_bbs_seq_range (e.g., "734 or 774")
  - excluded_bbs_seq (definite NOT)
  - basis (which markers + filename hints)
  - confidence_after_fingerprinting (still typically `needs_user_confirmation`)
  - proposed_bbs_seq (best single guess)
  - user_confirmation_needed (Y/N)

The user reviews this CSV + tells me corrections; I then re-run with user input.
"""
from __future__ import annotations

import csv
import hashlib
import re
from pathlib import Path

import openpyxl

SRC = Path("/home/jin/code/quant/research_input_data/정기변경")
OUT_DIR = Path("/home/jin/code/quant/data/acquired/bx01_kospi200_index_event_source_a0")
OUT = OUT_DIR / "class_b_mapping_proposal.csv"
TIER1 = OUT_DIR / "tier1_download_list.csv"

# Tier 1 target bbsSeq → reg_dt + cycle
TIER1_TARGETS = {
    "555": ("2018-05-24", "2018-06"),
    "613": ("2019-05-21", "2019-06"),
    "667": ("2020-05-27", "2020-06"),
    "696": ("2020-11-25", "2020-12"),
    "734": ("2021-05-25", "2021-06"),  # Class A covers this
    "774": ("2021-11-24", "2021-12"),
    "799": ("2022-05-24", "2022-06"),
    "861": ("2023-05-18", "2023-06"),
    "899": ("2023-11-23", "2023-12"),
    "939": ("2024-05-24", "2024-06"),
    "979": ("2024-11-21", "2024-12"),
    "1015": ("2025-05-27", "2025-06"),
    "1050": ("2025-11-18", "2025-12"),
    "1090": ("2026-05-22", "2026-06"),
}

# Marker stocks (ticker → (descriptive event, "added-on cycle", "stays in or after"))
# Inclusion events: if ticker IS in snapshot, snapshot is FROM that cycle onward.
INCLUSION_MARKERS = [
    ("298020", "효성티앤씨 added 2018-06", "2018-06"),
    ("298050", "효성첨단소재 added 2018-06", "2018-06"),
    ("302440", "SK바이오사이언스 added 2021-06", "2021-06"),
    ("259960", "크래프톤 added 2021-08", "2021-08"),  # special, might be in 2021-12 or later regular
    ("377300", "카카오뱅크 added 2021-08", "2021-08"),
    ("377740", "카카오페이 added 2021-11", "2021-11"),
    ("373220", "LG에너지솔루션 added 2022-01", "2022-01"),  # special
    ("450080", "에코프로머티 added 2023-12", "2023-12"),  # special
    ("095870", "포스코DX added 2024-01", "2024-01"),  # special
]

# Deletion events: if ticker IS NOT in snapshot, snapshot is FROM that cycle onward (after delete).
DELETION_MARKERS = [
    ("005180", "빙그레 deleted 2021-06", "2021-06"),
    ("005610", "SPC삼립 deleted 2021-06", "2021-06"),
    ("008350", "남선알미늄 deleted 2021-06", "2021-06"),
]


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def normalize_ticker(s: str) -> str | None:
    if not s:
        return None
    s = str(s).strip()
    m = re.match(r"^KR7(\d{6})\d{3}$", s)
    if m:
        return m.group(1)
    m = re.match(r"^(\d{6})$", s)
    if m:
        return m.group(1)
    m = re.match(r"^(\d{1,6})$", s)
    if m:
        return m.group(1).zfill(6)
    return None


def extract_kospi200_set(p: Path) -> set[str]:
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    out = set()
    for sn in wb.sheetnames:
        if "KOSPI 200" in sn and "구성종목" in sn:
            ws = wb[sn]
            for r in ws.iter_rows(values_only=True):
                if not r or not r[0]:
                    continue
                t = normalize_ticker(r[0])
                if t and t.isdigit():
                    out.add(t)
            break
    wb.close()
    return out


def bracket_cycle(kset: set[str]) -> tuple[str | None, str | None, list[str]]:
    """Use marker presence/absence to bracket cycle.
    Returns (earliest_cycle_at_least, evidence_lines)."""
    earliest = None
    evidence = []
    for t, desc, cycle in INCLUSION_MARKERS:
        present = t in kset
        evidence.append(f"{'✓' if present else '✗'} {t} {desc}")
        if present:
            if earliest is None or cycle > earliest:
                earliest = cycle
    for t, desc, cycle in DELETION_MARKERS:
        present = t in kset
        evidence.append(f"{'present:not-yet-deleted' if present else 'absent:after-delete'} {t} {desc}")
        if not present:
            if earliest is None or cycle > earliest:
                earliest = cycle
    return earliest, evidence


def main():
    files = sorted([p for p in SRC.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx"
                    and not str(p).endswith(":Zone.Identifier")])
    rows = []
    for p in files:
        # Skip Class A (has 변경내역 sheet)
        try:
            wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
            sheets = wb.sheetnames
            wb.close()
        except Exception as e:
            continue
        if any("변경내역" in s for s in sheets):
            continue  # Class A

        kset = extract_kospi200_set(p)
        earliest, evidence = bracket_cycle(kset)

        # Possible bbsSeq from marker bracket
        possible_seqs = [seq for seq, (_, cycle) in TIER1_TARGETS.items()
                          if earliest is None or cycle >= earliest]

        # Filename hints
        fn = p.name
        fn_year_hint = ""
        for token in ["18년", "19년", "20년", "21년", "22년", "23년", "24년", "25년", "26년",
                      "2018", "2019", "2020", "2021", "2022", "2023", "2024", "2025", "2026",
                      "21.06", "21.12", "26.6", "23.6", "23.12"]:
            if token in fn:
                fn_year_hint += token + " "

        has_valueup = any("밸류업" in s for s in sheets)
        if has_valueup:
            # Value-up index launched 2024-09 → likely 2024-12 (bbsSeq=979) or later (1015, 1050, 1090)
            possible_seqs = [s for s in possible_seqs if s in {"979", "1015", "1050", "1090"}]

        # Best single guess heuristic:
        # If only one candidate from above, that's the proposal.
        # Otherwise, propose the LATEST possible (most recent) since browser often downloads
        # in chronological order with newest first / first.
        # User confirmation always needed for Class B per Referee.
        if len(possible_seqs) == 1:
            proposed = possible_seqs[0]
            confidence = "single_candidate_from_markers"
        else:
            proposed = possible_seqs[0] if possible_seqs else ""
            confidence = "multiple_candidates_user_must_confirm"

        # Brief evidence summary (cut markers list)
        markers_short = [e for e in evidence if "✓" in e or "absent:after-delete" in e][:6]

        rows.append({
            "filename": p.name,
            "sha256_short": sha256_file(p)[:12],
            "bytes": p.stat().st_size,
            "kospi200_row_count": len(kset),
            "has_valueup_sheet": "yes" if has_valueup else "no",
            "earliest_cycle_at_least": earliest or "",
            "possible_bbs_seq": "|".join(possible_seqs),
            "possible_cycles": "|".join(f"{TIER1_TARGETS[s][1]}({s})" for s in possible_seqs),
            "filename_year_hints": fn_year_hint.strip(),
            "marker_evidence_short": " | ".join(markers_short),
            "proposed_bbs_seq": proposed,
            "proposed_cycle": TIER1_TARGETS[proposed][1] if proposed else "",
            "proposed_reg_dt": TIER1_TARGETS[proposed][0] if proposed else "",
            "confidence": confidence,
            "user_confirmation_needed": "YES — please confirm/correct proposed_bbs_seq",
        })

    cols = list(rows[0].keys())
    with OUT.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, lineterminator="\n")
        w.writeheader()
        w.writerows(rows)
    print(f"wrote {len(rows)} class-B rows to {OUT}\n")
    for r in rows:
        print(f"  {r['filename'][:50]}")
        print(f"    earliest>={r['earliest_cycle_at_least']:8s}  possible={r['possible_cycles']}")
        print(f"    proposed_bbs_seq={r['proposed_bbs_seq']} ({r['proposed_cycle']})  confidence={r['confidence']}")
        print()


if __name__ == "__main__":
    main()
