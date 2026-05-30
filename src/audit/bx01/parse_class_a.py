#!/usr/bin/env python3
"""BX01-attachment-parse: Class A parse (the 1 KRX xlsx with explicit 변경내역 sheet).

Per Referee clarification (2026-05-30):
- Class A = `21.06+KOSPI+200,+KOSDAQ+150,+KRX+300+공지파일.xlsx`
- High confidence, bbsSeq=734, review_cycle=2021-06.
- Parse KOSPI200 rows ONLY (not KOSDAQ150 / KRX300).
- effective_dt direct-from-file-only (NOT rebalance convention).
- Preserve-all + listing-name cross-check + caveats.

Output: events_class_a.csv with constituent-level rows for the 2021-06 review.
"""
from __future__ import annotations

import csv
import hashlib
from pathlib import Path

import openpyxl

SRC_FILE = Path(
    "/home/jin/code/quant/research_input_data/정기변경/"
    "21.06+KOSPI+200,+KOSDAQ+150,+KRX+300+공지파일.xlsx"
)
OUT_DIR = Path("/home/jin/code/quant/data/acquired/bx01_kospi200_index_event_source_a0")
OUT = OUT_DIR / "events_class_a.csv"

# Existing notice metadata for cross-reference
NOTICE_BBSSEQ = "734"
NOTICE_REG_DT = "2021-05-25"  # KRX REG_DT for bbsSeq=734
REVIEW_CYCLE = "2021-06"
SOURCE_NOTE = "user_supplied_from_KRX_MDCINFO005_bbsSeq=734"


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def parse_kospi200_change_detail(p: Path) -> list[dict]:
    """Extract KOSPI200 변경내역 rows from Class A xlsx."""
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    sheets = wb.sheetnames
    target_sn = next((s for s in sheets if "KOSPI 200" in s and "변경내역" in s), None)
    if target_sn is None:
        wb.close()
        raise RuntimeError(f"KOSPI200 변경내역 sheet not found; sheets={sheets}")

    ws = wb[target_sn]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Layout (from inspection):
    # row 0: ['KOSPI 200 지수 편입/제외종목', ...]
    # row 1: []
    # row 2: ['편입', '', '제외', '']
    # row 3: ['종목코드', '종목명', '종목코드', '종목명']  <- header
    # row 4+: [add_ticker, add_name, del_ticker, del_name]
    additions = []
    deletions = []
    # Find header row
    header_idx = None
    for i, r in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in r]
        if cells[:1] == ["종목코드"] and any("종목명" in c for c in cells[:4]):
            header_idx = i
            break
    if header_idx is None:
        raise RuntimeError(f"header row not found in 변경내역 sheet; first 6 rows: {rows[:6]}")

    # Each data row: [add_code, add_name, del_code, del_name]
    for r in rows[header_idx + 1 :]:
        cells = [c if c is not None else "" for c in r]
        if len(cells) < 4:
            continue
        a_code, a_name, d_code, d_name = cells[0], cells[1], cells[2], cells[3]
        if a_code and str(a_code).strip():
            additions.append({"ticker": str(a_code).strip(), "name": str(a_name).strip()})
        if d_code and str(d_code).strip():
            deletions.append({"ticker": str(d_code).strip(), "name": str(d_name).strip()})

    return [
        *[{"side": "addition", **a} for a in additions],
        *[{"side": "deletion", **d} for d in deletions],
    ]


def main():
    if not SRC_FILE.exists():
        raise SystemExit(f"missing: {SRC_FILE}")

    sha = sha256_file(SRC_FILE)
    print(f"source: {SRC_FILE.name}")
    print(f"sha256: {sha}")
    print(f"bytes:  {SRC_FILE.stat().st_size}")
    print(f"bbsSeq: {NOTICE_BBSSEQ}  cycle: {REVIEW_CYCLE}  reg_dt: {NOTICE_REG_DT}\n")

    detail_rows = parse_kospi200_change_detail(SRC_FILE)
    n_add = sum(1 for r in detail_rows if r["side"] == "addition")
    n_del = sum(1 for r in detail_rows if r["side"] == "deletion")
    print(f"parsed KOSPI200 변경내역: {n_add} additions + {n_del} deletions")
    for r in detail_rows:
        print(f"  {r['side']:8s} {r['ticker']:>8s} {r['name']}")

    # Write events_class_a.csv
    cols = [
        "event_id", "review_cycle", "announcement_dt", "announcement_time", "effective_dt",
        "index_name", "side", "ticker", "company_name",
        "source_class", "source_id_or_url", "source_file", "source_file_sha256", "source_note",
        "parse_status", "caveat",
    ]
    out_rows = []
    for i, r in enumerate(detail_rows, 1):
        eid = f"BX01-PARSE-{NOTICE_BBSSEQ}-{r['side'][:3].upper()}-{i:02d}"
        out_rows.append({
            "event_id": eid,
            "review_cycle": REVIEW_CYCLE,
            "announcement_dt": NOTICE_REG_DT,
            "announcement_time": "",  # not in attachment; not in metadata at row-level
            "effective_dt": "",  # NOT in attachment per inspection; left blank per Referee rule
            "index_name": "KOSPI 200",
            "side": r["side"],
            "ticker": r["ticker"],
            "company_name": r["name"],
            "source_class": "A_KRX_xlsx_with_change_detail",
            "source_id_or_url": (
                f"KRX MDCINFO005 bbsSeq={NOTICE_BBSSEQ} "
                f"(https://data.krx.co.kr/contents/MDC/COMS/board/MDCCOMS010_S2.cmd"
                f"?boardId=MDCINFO005&cmBbsId=MKD01040000&bbsSeq={NOTICE_BBSSEQ})"
            ),
            "source_file": SRC_FILE.name,
            "source_file_sha256": sha,
            "source_note": SOURCE_NOTE,
            "parse_status": "parsed_direct_from_kospi200_변경내역_sheet",
            "caveat": (
                "ticker is 6-digit KRX format from attachment; ISIN cross-check not yet "
                "performed; listing-name cross-check against listing-universe-A0 outputs not "
                "yet performed; effective_dt blank because not present in attachment "
                "(Referee rule: no rule/convention fill)"
            ),
        })

    with OUT.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, lineterminator="\n")
        w.writeheader()
        w.writerows(out_rows)
    print(f"\nwrote {len(out_rows)} rows to {OUT}")


if __name__ == "__main__":
    main()
