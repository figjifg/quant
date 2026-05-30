#!/usr/bin/env python3
"""BX01-attachment-parse: build coverage matrix mapping intake files → Tier 1 bbsSeq.

Per Referee clarification (2026-05-30): build coverage matrix FIRST. Auto fingerprint
allowed only as SUPPORTING evidence. High-confidence mapping requires multiple
corroborating signals; ambiguous files = `cycle_unresolved` / `needs_user_cycle_confirmation`,
NO parse / derivation.

Read-only over research_input_data/정기변경/. Writes to data/acquired/.../coverage_matrix.csv.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
from pathlib import Path
from typing import Optional

import openpyxl

SRC = Path("/home/jin/code/quant/research_input_data/정기변경")
OUT_DIR = Path("/home/jin/code/quant/data/acquired/bx01_kospi200_index_event_source_a0")
OUT = OUT_DIR / "coverage_matrix.csv"
TIER1 = OUT_DIR / "tier1_download_list.csv"

# Marker stocks for cycle bracketing — based on PUBLIC KRX events:
# (ticker, marker_event, applies_to_cycle_or_later)
#
# This is SUPPORTING EVIDENCE per Referee — never high-confidence alone. We never
# infer a cycle from a single marker; we use the FULL marker set to bracket.
# Tickers used in standard KRX 6-digit format (some xlsx use KR7XXXXXX002 ISIN format).
MARKERS = [
    # Inclusion events — once present, must be in the listed cycle or later
    ("298020", "효성티앤씨 added 2018-06 split", "2018-06"),
    ("298050", "효성첨단소재 added 2018-06 split", "2018-06"),
    ("302440", "SK바이오사이언스 added 2021-06 review", "2021-06"),
    ("259960", "크래프톤 added 2021-08 special", "2021-08"),
    ("377300", "카카오뱅크 added 2021-08 special", "2021-08"),
    ("377740", "카카오페이 added 2021-11", "2021-11"),
    ("373220", "LG에너지솔루션 added 2022-01 special", "2022-01"),
    ("450080", "에코프로머티 added 2023-12 special", "2023-12"),
    ("095870", "포스코DX added 2024-01 special", "2024-01"),
    # Deletion events — once absent, must be after this cycle
    ("005930", "삼성전자 always present", "ALL"),  # sanity
]


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def normalize_ticker(s: str) -> Optional[str]:
    if not s:
        return None
    s = str(s).strip()
    # ISIN format KR7XXXXXX00Y → 6-digit XXXXXX
    m = re.match(r"^KR7(\d{6})\d{3}$", s)
    if m:
        return m.group(1)
    # short 6-digit
    m = re.match(r"^(\d{6})$", s)
    if m:
        return m.group(1)
    # leading-zero stripped
    m = re.match(r"^(\d{1,6})$", s)
    if m:
        return m.group(1).zfill(6)
    return None


def extract_kospi200_set(p: Path) -> tuple[set[str], Optional[str], list[str]]:
    """Return (set of 6-digit tickers, header_text_first_sheet, sheet_names)."""
    try:
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        sheets = wb.sheetnames
        kospi_set = set()
        header_text = None
        for sn in sheets:
            if "KOSPI 200" not in sn:
                continue
            if "구성종목" not in sn:
                continue
            ws = wb[sn]
            for r in ws.iter_rows(values_only=True):
                if not r:
                    continue
                if header_text is None and r[0] and isinstance(r[0], str) and "KOSPI 200" in r[0]:
                    header_text = r[0]
                if not r[0]:
                    continue
                t = normalize_ticker(r[0])
                if t and t.isdigit():
                    kospi_set.add(t)
            break  # first KOSPI200 구성종목 sheet only
        wb.close()
        return kospi_set, header_text, sheets
    except Exception as e:
        return set(), f"ERR: {e}", []


def fingerprint_cycle_bracket(kospi_set: set[str]) -> dict:
    """Return cycle bracketing evidence from marker presence/absence."""
    earliest_lower_bound = None  # cycle this snapshot is AT LEAST
    marker_evidence = []
    for ticker, desc, cycle in MARKERS:
        present = ticker in kospi_set
        marker_evidence.append({
            "ticker": ticker, "desc": desc, "cycle_marker": cycle, "present": present,
        })
        if present and cycle != "ALL":
            if earliest_lower_bound is None or cycle > earliest_lower_bound:
                earliest_lower_bound = cycle
    return {
        "earliest_lower_bound": earliest_lower_bound,
        "marker_evidence": marker_evidence,
    }


def load_tier1_targets() -> list[dict]:
    return list(csv.DictReader(TIER1.open(encoding="utf-8")))


def main():
    targets = load_tier1_targets()
    target_by_seq = {t["bbs_seq"]: t for t in targets}
    print(f"Tier 1 targets: {len(targets)} (bbsSeq {sorted(target_by_seq)})")

    files = sorted([p for p in SRC.iterdir() if p.is_file() and p.suffix.lower() in (".xlsx",)
                    and not str(p).endswith(":Zone.Identifier")])
    print(f"xlsx files to evaluate: {len(files)}\n")

    rows = []
    for p in files:
        kset, hdr, sheets = extract_kospi200_set(p)
        bracket = fingerprint_cycle_bracket(kset) if kset else None
        sha = sha256_file(p)
        has_change_detail = any("변경내역" in s or "편입/제외" in s for s in sheets)
        has_korea_valueup = any("밸류업" in s for s in sheets)

        # Filename-based cycle hint (no high-confidence — only hint)
        fn = p.name
        fn_hints = []
        for token in ["18년", "19년", "20년", "21년", "22년", "23년", "24년", "25년", "26년",
                      "2018", "2019", "2020", "2021", "2022", "2023", "2024", "2025", "2026",
                      "06월", "12월", "6월", "12월", "26.6", "26.12", "21.06", "21.12"]:
            if token in fn:
                fn_hints.append(token)

        # Confidence assessment
        candidate_bbs_seq = ""
        candidate_cycle = ""
        confidence = "unresolved"
        basis = []

        # Class A: change-detail
        if has_change_detail:
            source_class = "A_KRX_xlsx_with_change_detail"
            if "21.06" in fn or "21년 06월" in fn:
                candidate_bbs_seq = "734"
                candidate_cycle = "2021-06"
                confidence = "high"
                basis.append("filename token 21.06 matches 2021-06 review (bbsSeq=734)")
            else:
                confidence = "needs_user_cycle_confirmation"
                basis.append("change-detail format but filename does not unambiguously bind to bbsSeq")
        elif has_korea_valueup:
            source_class = "B_KRX_xlsx_snapshot_only_with_valueup_sheet"
            # Value-up index was launched 2024-09 → only 2024-12+ snapshots can contain it.
            # Need to check if it's exactly 2024-12, 2025-06, 2025-12, or 2026-06.
            confidence = "needs_user_cycle_confirmation"
            basis.append("코리아 밸류업 sheet present → 2024-12+ era; needs filename or content marker for specific cycle")
            if bracket and bracket["earliest_lower_bound"]:
                basis.append(f"earliest_lower_bound from markers: {bracket['earliest_lower_bound']}")
        else:
            source_class = "B_KRX_xlsx_snapshot_only"
            # Try marker bracketing
            if bracket:
                lb = bracket["earliest_lower_bound"]
                if lb:
                    basis.append(f"earliest_lower_bound from markers: {lb} (at-least, not exact)")
            confidence = "needs_user_cycle_confirmation"
            basis.append("snapshot-only file; cycle identification requires direct match against KRX notice REG_DT (not just markers)")
            # Filename hint?
            if fn_hints:
                basis.append(f"filename hints: {fn_hints}")

        # KOSPI 200 size sanity
        kosize = len(kset)
        if kset and kosize != 200:
            basis.append(f"KOSPI200 row count = {kosize} (expected 200 — unusual)")

        # Build row
        rows.append({
            "filename": p.name,
            "sha256": sha,
            "source_path": str(p),
            "source_class": source_class,
            "has_change_detail_sheet": "yes" if has_change_detail else "no",
            "has_korea_valueup_sheet": "yes" if has_korea_valueup else "no",
            "sheets": "|".join(sheets),
            "kospi200_row_count": kosize,
            "snapshot_only": "yes" if not has_change_detail else "no",
            "candidate_bbs_seq": candidate_bbs_seq,
            "candidate_review_cycle": candidate_cycle,
            "mapping_confidence": confidence,
            "basis": "; ".join(basis) if basis else "",
            "filename_hints": "|".join(fn_hints) if fn_hints else "",
            "markers_summary": ("|".join(f"{m['ticker']}={'1' if m['present'] else '0'}"
                                          for m in (bracket["marker_evidence"] if bracket else []))),
            "parse_eligibility": ("primary_change_detail_parse" if has_change_detail and confidence == "high"
                                   else "snapshot_table_only_if_confirmed" if not has_change_detail
                                   else "needs_user_cycle_confirmation"),
            "next_action": ("parse_class_A_kospi200_rows" if has_change_detail and confidence == "high"
                            else "request_user_cycle_confirmation"),
            "caveat": ("change_detail_format" if has_change_detail
                       else "snapshot_only_format__cycle_identification_required_per_Referee_clarification"),
        })

    # Also add non-xlsx files (hwp + pdf) with deferred status
    for p in sorted([p for p in SRC.iterdir() if p.is_file() and p.suffix.lower() in (".hwp", ".pdf")
                     and not str(p).endswith(":Zone.Identifier")]):
        ext = p.suffix.lower()
        sha = sha256_file(p)
        if ext == ".hwp":
            cls = "C_KRX_hwp_press_release_body_text"
            note = "deferred_per_Referee — hwp parsing not in current phase scope; evidence-only"
        else:
            cls = "D_secondary_broker_pdf_user_supplied"
            note = "deferred_per_Referee — broker publication, source authority differs from KRX direct; secondary cross-check only, not record-of-record"
        rows.append({
            "filename": p.name,
            "sha256": sha,
            "source_path": str(p),
            "source_class": cls,
            "has_change_detail_sheet": "n/a",
            "has_korea_valueup_sheet": "n/a",
            "sheets": "-",
            "kospi200_row_count": "",
            "snapshot_only": "n/a",
            "candidate_bbs_seq": "",
            "candidate_review_cycle": "",
            "mapping_confidence": "deferred_class",
            "basis": "",
            "filename_hints": "",
            "markers_summary": "",
            "parse_eligibility": "deferred",
            "next_action": "evidence_only_inventory",
            "caveat": note,
        })

    rows.sort(key=lambda r: r["source_class"] + "_" + r["filename"])

    cols = list(rows[0].keys())
    with OUT.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, lineterminator="\n")
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print(f"wrote {len(rows)} rows to {OUT}\n")

    # Summary by class
    from collections import Counter
    print("by source_class:")
    for k, v in Counter(r["source_class"] for r in rows).items():
        print(f"  {v}  {k}")
    print()
    print("by mapping_confidence:")
    for k, v in Counter(r["mapping_confidence"] for r in rows).items():
        print(f"  {v}  {k}")
    print()
    print("by next_action:")
    for k, v in Counter(r["next_action"] for r in rows).items():
        print(f"  {v}  {k}")


if __name__ == "__main__":
    main()
